import streamlit as st
import yfinance as yf
import pandas as pd
import plotly.express as px
from datetime import datetime
from pycoingecko import CoinGeckoAPI
import finnhub
import openpyxl  # For Excel file handling

# Function to load the workbook and create a new one if it doesn't exist
def load_workbook(filename="users.xlsx"):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        # Create a new workbook and add headers if the file does not exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Password"])  # Adding headers for username and password columns
        workbook.save(filename)
    return workbook

# Function to handle both Log In and Sign Up on the same page
def authentication_page(filename="users.xlsx"):
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Create two columns for Log In and Sign Up
    col1, col2 = st.columns(2)

    # Log In Section
    with col1:
        st.subheader("Log In")
        login_username = st.text_input("Username", key="login_username")
        login_password = st.text_input("Password", type="password", key="login_password")
        if st.button("Log In"):
            # Check if the username and password match
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0] == login_username and row[1] == login_password:
                    st.session_state["logged_in"] = True
                    st.session_state["page"] = "dashboard"
                    st.success("Logged in successfully!")
                    st.rerun()
                    return
            st.error("Invalid username or password.")

    # Sign Up Section
    with col2:
        st.subheader("Sign Up")
        signup_username = st.text_input("New Username", key="signup_username")
        signup_password = st.text_input("New Password", type="password", key="signup_password")
        if st.button("Sign Up"):
            # Check if the username already exists
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                if row[0] == signup_username:
                    st.error("Username already exists!")
                    return False
            # Save the new username and password
            sheet.append([signup_username, signup_password])
            workbook.save(filename)
            st.success("Sign-up successful! You can now log in.")
            st.session_state["page"] = "login"
            st.rerun()
            return
        
    return False


# Displays stock news
def fetch_stock_news_finnhub(ticker):
    """Fetches recent stock related news articles in the top 5 using Finnhub."""
    finnhub_client = finnhub.Client(api_key="ctcvpg1r01qlc0uvnvngctcvpg1r01qlc0uvnvo0")  # Custom API key we got from Finnhub website.
    try:
        # Fetch news
        news = finnhub_client.company_news(ticker, _from='2023-01-01', to=datetime.today().strftime('%Y-%m-%d'))
        # Return the top 5 news articles
        return news[:5]
    except Exception as e:
        st.error(f"Error fetching news for {ticker}: {e}")
        return []
def display_stock_news(news):
    """Display news articles for a stock ticker."""
    st.subheader("Latest News")
    if not news:
        st.info("No news available.")
    else:
        for article in news:
            st.markdown(f"### [{article['headline']}]({article['url']})")
            st.markdown(f"**Source:** {article['source']} | **Date:** {article['datetime']}")
            st.markdown(f"> {article['summary']}")
            st.markdown("---")

# Fetch live stock data
def fetch_live_stock_data(tickers):
    """Fetch live stock data for multiple tickers or fallback to last closing price."""
    data = []  # List to store stock data
    for ticker in tickers:
        try:
            stock = yf.Ticker(ticker)
            info = stock.info  # Fetch live data

            # Fetch the last close as fallback
            last_close = stock.history(period="1d").iloc[-1]["Close"] if not stock.history(period="1d").empty else None

            # Handle missing data
            price = info.get("regularMarketPrice", last_close)
            day_high = info.get("dayHigh", 0)
            day_low = info.get("dayLow", 0)
            volume = info.get("volume", 0)
            
            # Add stock data
            if price is not None:
                data.append({
                    "Ticker": ticker,
                    "Price": round(price, 2),
                    "Day High": round(day_high, 2),
                    "Day Low": round(day_low, 2),
                    "Volume": int(volume),
                })
            else:
                st.warning(f"No live or closing data available for {ticker}.")
        except Exception as e:
            st.error(f"Error fetching data for {ticker}: {e}")
    return pd.DataFrame(data)


# Fetch historical stock data
def fetch_stock_data(tickers, start_date, end_date):
    data = {}  # Dictionary to store stock data
    for ticker in tickers:
        try:
            stock = yf.Ticker(ticker)
            df = stock.history(start=start_date, end=end_date)
            if not df.empty:
                df['Ticker'] = ticker
                data[ticker] = df
            else:
                st.warning(f"No data available for {ticker} in the selected date range.")
        except Exception as e:
            st.error(f"Error fetching data for {ticker}: {e}")
    return data

# Fetch historical cryptocurrency data
def fetch_crypto_data(crypto_ids, start_date, end_date):
    cg = CoinGeckoAPI()
    data = {}
    for crypto in crypto_ids:
        try:
            start_timestamp = int(pd.Timestamp(start_date).timestamp())
            end_timestamp = int(pd.Timestamp(end_date).timestamp())

            market_data = cg.get_coin_market_chart_range_by_id(
                id=crypto,
                vs_currency='usd',
                from_timestamp=start_timestamp,
                to_timestamp=end_timestamp
            )

            if 'prices' in market_data:
                prices = market_data['prices']
                df = pd.DataFrame(prices, columns=['timestamp', 'Close'])
                df['Date'] = pd.to_datetime(df['timestamp'], unit='ms')
                df['Close'] = df['Close'].astype(float)
                df['Ticker'] = crypto.upper()
                data[crypto] = df[['Date', 'Close', 'Ticker']]
            else:
                st.warning(f"No data available for {crypto}.")
        except Exception as e:
            st.error(f"Error fetching data for {crypto}: {e}")
    return data

# Display a summary of the latest data
def display_stock_summary(data):
    summary = []
    for ticker, df in data.items():
        if not df.empty:
            latest = df.iloc[-1]
            summary.append({
                "Ticker": ticker,
                "Open": round(latest.get("Open", 0), 2),
                "High": round(latest.get("High", 0), 2),
                "Low": round(latest.get("Low", 0), 2),
                "Close": round(latest["Close"], 2),
                "Volume": int(latest.get("Volume", 0)),
            })
    summary_df = pd.DataFrame(summary)
    if not summary_df.empty:
        st.subheader("Summary")
        st.dataframe(summary_df)

# Plot closing prices for stocks and cryptos
def plot_stock_data(data):
    combined_data = pd.concat(data.values())
    combined_data.reset_index(inplace=True)

    fig = px.line(
        combined_data,
        x='Date',
        y='Close',
        color='Ticker',
        title="Closing Price Trends"
    )
    st.plotly_chart(fig)
    
# Main dashboard
def display_dashboard():
    """Displays the main Pynance dashboard."""
    st.markdown("<h1 style='color:gold; font-family:sans-serif;'>Pynance $ ;)</h1>", unsafe_allow_html=True)
    tickers_input = st.text_input(
        "Enter stock/crypto tickers or IDs (comma-separated)",
        key="tickers_input",
        help="Press Enter after typing to analyze data."
    )
    live_data = st.checkbox("Fetch Live Data")
    crypto_mode = st.checkbox("Include Cryptocurrencies")
    start_date = st.date_input("Start date", pd.to_datetime("2022-01-01"))
    end_date = st.date_input("End date", pd.to_datetime(datetime.today().strftime('%Y-%m-%d')))

    analyze_button = st.button("Analyze Data")
    if analyze_button:
        tickers = [ticker.strip().upper() for ticker in tickers_input.split(",") if ticker.strip()]
        stocks = []
        crypto_ids = []

        if crypto_mode:
            for ticker in tickers:
                if ticker.lower() in ['bitcoin', 'ethereum', 'dogecoin', 'cardano', 'litecoin']:
                    crypto_ids.append(ticker.lower())
                else:
                    stocks.append(ticker)
        else:
            stocks = tickers

        if live_data:
            stock_data = fetch_live_stock_data(stocks) if stocks else pd.DataFrame()
            if not stock_data.empty:
                st.subheader("Live Stock Data")
                st.dataframe(stock_data)
        else:
            stock_data = fetch_stock_data(stocks, start_date, end_date) if stocks else {}
            crypto_data = fetch_crypto_data(crypto_ids, start_date, end_date) if crypto_ids else {}

            all_data = {**stock_data, **crypto_data}

            display_stock_summary(all_data)
            if all_data:
                plot_stock_data(all_data)

        for ticker in tickers:
            st.subheader(f"News for {ticker}")
            news = fetch_stock_news_finnhub(ticker)
            display_stock_news(news)


# Main app logic
def main():
     # Initialize session states
    if "logged_in" not in st.session_state:
        st.session_state["logged_in"] = False
    if "page" not in st.session_state:
        st.session_state["page"] = "login"

    # Page routing
    if st.session_state["page"] == "login" and not st.session_state["logged_in"]:
        authentication_page()
    elif st.session_state["page"] == "dashboard" and st.session_state["logged_in"]:
        display_dashboard()
    else:
        st.session_state["page"] = "login"
        st.rerun()



if __name__ == "__main__":
    main()
